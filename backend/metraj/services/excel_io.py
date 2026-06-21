import io
from decimal import Decimal, InvalidOperation

from openpyxl import Workbook, load_workbook

from metraj.models import MetrajCategory, MetrajItem

TEMPLATE_HEADERS = [
    "Kategori",
    "Açıklama",
    "Birim",
    "Miktar",
    "Birim Fiyat",
    "İlerleme %",
    "Notlar",
]

CATEGORY_ALIASES = {
    "beton": "beton",
    "demir": "demir",
    "donati": "demir",
    "donatı": "demir",
    "siva": "siva",
    "sıva": "siva",
    "kalip": "kalip",
    "kalıp": "kalip",
    "boya": "boya",
    "izolasyon": "izolasyon",
}


def _parse_decimal(value, default=Decimal("0")):
    if value is None or value == "":
        return default
    try:
        return Decimal(str(value).replace(",", "."))
    except (InvalidOperation, ValueError):
        return default


def _parse_int(value, default=0):
    try:
        val = int(float(str(value).replace(",", ".")))
        return max(0, min(100, val))
    except (TypeError, ValueError):
        return default


def _resolve_category(raw: str) -> MetrajCategory | None:
    key = (raw or "").strip().lower()
    slug = CATEGORY_ALIASES.get(key, key)
    return MetrajCategory.objects.filter(slug=slug).first()


def build_template_workbook() -> io.BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Metraj"
    sheet.append(TEMPLATE_HEADERS)
    sheet.append(["beton", "Temel betonu", "m3", 120, 850, 45, ""])
    sheet.append(["demir", "Temel donatısı", "ton", 12.5, 0, 30, ""])
    sheet.append(["siva", "Dış cephe sıvası", "m2", 450, 120, 10, ""])
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def import_metraj_from_workbook(site, file_obj) -> list[MetrajItem]:
    workbook = load_workbook(file_obj, read_only=True, data_only=True)
    sheet = workbook.active
    rows = [
        row
        for row in sheet.iter_rows(values_only=True)
        if row and any(cell is not None and str(cell).strip() for cell in row)
    ]
    if not rows:
        return []

    start = 0
    if str(rows[0][0] or "").strip().lower().startswith("kategori"):
        start = 1

    MetrajItem.objects.filter(site=site).delete()
    created: list[MetrajItem] = []

    for row in rows[start:]:
        cells = list(row) + [None] * (7 - len(row))
        category = _resolve_category(str(cells[0] or ""))
        if not category:
            continue

        unit_raw = str(cells[2] or category.default_unit).strip().lower()
        valid_units = {choice[0] for choice in MetrajItem.Unit.choices}
        unit = unit_raw if unit_raw in valid_units else category.default_unit

        price_raw = cells[4]
        unit_price = None
        if price_raw not in (None, ""):
            unit_price = _parse_decimal(price_raw)

        item = MetrajItem.objects.create(
            site=site,
            category=category,
            description=str(cells[1] or "").strip() or category.name,
            unit=unit,
            quantity=_parse_decimal(cells[3]),
            unit_price=unit_price,
            completion_percent=_parse_int(cells[5]),
            notes=str(cells[6] or "").strip(),
        )
        created.append(item)

    return created


def export_metraj_workbook(site) -> io.BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Metraj"
    sheet.append(TEMPLATE_HEADERS)
    for item in site.metraj_items.select_related("category").order_by("category__sort_order", "id"):
        sheet.append(
            [
                item.category.slug,
                item.description,
                item.unit,
                float(item.quantity),
                float(item.unit_price) if item.unit_price is not None else "",
                item.completion_percent,
                item.notes,
            ]
        )
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
