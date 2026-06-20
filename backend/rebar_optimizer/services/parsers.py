"""Aşama 1 — Veri Ayıklama (Parsing Engine).

DXF, PDF ve XLSX kaynaklarından donatı verisini ayıklayıp standart JSON'a çevirir.

Standart çıktı formatı (her satır):
    {
        "rebar_diameter": 16,      # mm
        "total_length": 4.5,       # metre (tek parça boyu)
        "quantity": 10,            # adet
        "element_ref": "K-101"     # eleman referansı (opsiyonel)
    }
"""

from __future__ import annotations

import re
from typing import IO, Any

# Donatı notasyonu: "Ø16", "T16", "D16", "fi16", "Φ16", "⌀16" + boy/adet
_DIAMETER_RE = re.compile(r"(?:Ø|Φ|⌀|T|D|fi|FI|q|Q)\s?0?(\d{1,2})\b", re.IGNORECASE)
_COUNT_DIA_RE = re.compile(
    r"(\d{1,3})\s?(?:Ø|Φ|⌀|T|D|fi|FI|q|Q)\s?0?(\d{1,2})\b", re.IGNORECASE
)
_LENGTH_RE = re.compile(
    r"(?:L\s?=?\s?|boy\s?:?\s?|uzunluk\s?:?\s?|length\s?:?\s?)(\d+(?:[.,]\d+)?)",
    re.IGNORECASE,
)
_QTY_RE = re.compile(r"(\d+)\s?(?:adet|ad\.?|x|pcs|nos?)\b", re.IGNORECASE)


def _is_plausible_length(value: float) -> bool:
    """Boy değeri makul mü? (50–1300 cm veya 0.5–13 m aralığı)."""
    return (50 <= value <= 1300) or (0.5 <= value <= 13)


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    try:
        return float(str(value).replace(",", "."))
    except (TypeError, ValueError):
        return None


def _to_int(value: Any) -> int | None:
    parsed = _to_float(value)
    if parsed is None:
        return None
    return int(round(parsed))


def _normalize_length_to_m(length: float) -> float:
    """100'den büyük değerleri cm kabul edip metreye çevirir (ör. 450 -> 4.5)."""
    if length > 100:
        return round(length / 100.0, 3)
    return round(length, 3)


def _make_row(diameter: Any, length: Any, quantity: Any, ref: str = "") -> dict | None:
    diameter_int = _to_int(diameter)
    length_float = _to_float(length)
    if not diameter_int or diameter_int <= 0:
        return None
    if not length_float or length_float <= 0:
        return None
    quantity_int = _to_int(quantity) or 1
    if quantity_int <= 0:
        quantity_int = 1
    return {
        "rebar_diameter": diameter_int,
        "total_length": _normalize_length_to_m(length_float),
        "quantity": quantity_int,
        "element_ref": (ref or "").strip()[:64],
    }


def parse_dxf(file: IO[bytes] | str) -> list[dict]:
    """DXF dosyasındaki donatı (rebar) katmanı metinlerinden veri ayıklar."""
    import ezdxf

    if hasattr(file, "read"):
        import io

        raw = file.read()
        text_stream = io.StringIO(raw.decode("utf-8", errors="ignore") if isinstance(raw, bytes) else raw)
        doc = ezdxf.read(text_stream)
    else:
        doc = ezdxf.readfile(file)

    msp = doc.modelspace()
    rebar_layers = {"rebar", "donati", "donatı", "demir", "reinforcement", "s-rebar"}
    rows: list[dict] = []

    for entity in msp:
        if entity.dxftype() not in {"TEXT", "MTEXT"}:
            continue
        layer = str(getattr(entity.dxf, "layer", "")).strip().lower()
        if rebar_layers and layer not in rebar_layers:
            # Katman eşleşmese de donatı notasyonu içeren metinleri değerlendir
            pass

        content = entity.plain_text() if entity.dxftype() == "MTEXT" else entity.dxf.text
        row = _parse_text_fragment(content)
        if row:
            rows.append(row)

    return _merge_rows(rows)


def _parse_text_fragment(text: str) -> dict | None:
    if not text:
        return None
    text = text.strip()
    diameter_match = _DIAMETER_RE.search(text)
    if not diameter_match:
        return None
    diameter = diameter_match.group(1)

    # Adet: "10Ø16" gibi öndeki sayı ya da "x10 / 10 adet"
    quantity_value = 1
    count_dia = _COUNT_DIA_RE.search(text)
    if count_dia and count_dia.group(2) == diameter:
        quantity_value = _to_int(count_dia.group(1)) or 1
    qty_match = _QTY_RE.search(text)
    if qty_match:
        quantity_value = _to_int(qty_match.group(1)) or quantity_value

    # Boy: önce "L=..." kalıbı, sonra çaptan sonraki makul sayı
    length_value: float | None = None
    length_match = _LENGTH_RE.search(text)
    if length_match:
        length_value = _to_float(length_match.group(1))

    if length_value is None:
        for number in re.findall(r"\d+(?:[.,]\d+)?", text[diameter_match.end():]):
            candidate = _to_float(number)
            if candidate is not None and _is_plausible_length(candidate):
                length_value = candidate
                break

    if length_value is None:
        return None
    return _make_row(diameter, length_value, quantity_value)


def parse_pdf(file: IO[bytes] | str) -> list[dict]:
    """PDF içindeki tablolardan ve metinden donatı verisini ayıklar."""
    import pdfplumber

    rows: list[dict] = []
    with pdfplumber.open(file) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables() or []:
                rows.extend(_parse_table(table))

            text = page.extract_text() or ""
            for line in text.splitlines():
                row = _parse_text_fragment(line)
                if row:
                    rows.append(row)

    return _merge_rows(rows)


def _parse_table(table: list[list[Any]]) -> list[dict]:
    if not table:
        return []

    header = [str(cell or "").strip().lower() for cell in table[0]]
    col = {"diameter": None, "length": None, "quantity": None, "ref": None}
    for idx, name in enumerate(header):
        if any(key in name for key in ("çap", "cap", "diameter", "ø", "diam")):
            col["diameter"] = idx
        elif any(key in name for key in ("boy", "length", "uzunluk")):
            col["length"] = idx
        elif any(key in name for key in ("adet", "quantity", "qty", "miktar")):
            col["quantity"] = idx
        elif any(key in name for key in ("eleman", "ref", "poz", "element")):
            col["ref"] = idx

    if col["diameter"] is None or col["length"] is None:
        return []

    rows: list[dict] = []
    for raw in table[1:]:
        def cell(index: int | None) -> Any:
            if index is None or index >= len(raw):
                return None
            return raw[index]

        ref = str(cell(col["ref"]) or "").strip()
        row = _make_row(cell(col["diameter"]), cell(col["length"]), cell(col["quantity"]), ref)
        if row:
            rows.append(row)
    return rows


def parse_xlsx(file: IO[bytes] | str) -> list[dict]:
    """MetrajX XLSX şablonundan donatı verisini okur.

    Beklenen sütun başlıkları (ilk satır): çap | boy | adet | eleman
    """
    from openpyxl import load_workbook

    workbook = load_workbook(file, read_only=True, data_only=True)
    sheet = workbook.active
    rows: list[dict] = []

    header_seen = False
    for excel_row in sheet.iter_rows(values_only=True):
        if excel_row is None:
            continue
        cells = list(excel_row)
        if not header_seen:
            header_seen = True
            first = str(cells[0] or "").strip().lower()
            if any(key in first for key in ("çap", "cap", "diameter", "ø")):
                continue  # başlık satırını atla
        if all(cell is None for cell in cells):
            continue
        diameter = cells[0] if len(cells) > 0 else None
        length = cells[1] if len(cells) > 1 else None
        quantity = cells[2] if len(cells) > 2 else None
        ref = str(cells[3]).strip() if len(cells) > 3 and cells[3] is not None else ""
        row = _make_row(diameter, length, quantity, ref)
        if row:
            rows.append(row)

    workbook.close()
    return _merge_rows(rows)


def _merge_rows(rows: list[dict]) -> list[dict]:
    """Aynı (çap, boy, eleman) satırlarını adet toplayarak birleştirir."""
    merged: dict[tuple, dict] = {}
    for row in rows:
        key = (row["rebar_diameter"], row["total_length"], row["element_ref"])
        if key in merged:
            merged[key]["quantity"] += row["quantity"]
        else:
            merged[key] = dict(row)
    return list(merged.values())


def parse_file(file: IO[bytes] | str, filename: str) -> list[dict]:
    """Uzantıya göre uygun parser'ı seçer."""
    name = filename.lower()
    if name.endswith(".dxf"):
        return parse_dxf(file)
    if name.endswith(".pdf"):
        return parse_pdf(file)
    if name.endswith((".xlsx", ".xlsm")):
        return parse_xlsx(file)
    raise ValueError(f"Desteklenmeyen dosya türü: {filename}. DXF, PDF veya XLSX yükleyin.")
