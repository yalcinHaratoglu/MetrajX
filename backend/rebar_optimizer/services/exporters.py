"""Aşama 3 — Raporlama (Export Engine).

Optimizasyon sonucunu iki formatta sunar:
    - Excel (openpyxl): çap bazlı metraj, çubuk ve fire özeti + kesim planı
    - PDF (reportlab): şantiyede kullanılabilir görsel "Kesim Çizelgesi"
Ayrıca kullanıcının doldurabileceği boş XLSX şablonu üretir.
"""

from __future__ import annotations

import io


TEMPLATE_HEADERS = ["Çap (mm)", "Boy (m)", "Adet", "Eleman No"]
TEMPLATE_EXAMPLES = [
    [16, 4.5, 10, "K-101"],
    [12, 3.0, 24, "K-102"],
    [8, 2.4, 50, "D-201"],
]


def _group_bars(bars: list[dict]) -> list[dict]:
    """Aynı kesim desenine sahip çubukları gruplar: {bar, count}.

    Aynı (boy + eleman no) sırası ve fire değerine sahip çubuklar tek satırda
    "×N" olarak gösterilir; böylece çok sayıda özdeş çubuk tekrar etmez.
    """
    groups: list[dict] = []
    index: dict[tuple, int] = {}
    for bar in bars:
        signature = (
            tuple((cut["length"], cut.get("element_ref", "")) for cut in bar["cuts"]),
            bar["waste_m"],
        )
        if signature in index:
            groups[index[signature]]["count"] += 1
        else:
            index[signature] = len(groups)
            groups.append({"bar": bar, "count": 1})
    return groups


def _cuts_label(cuts: list[dict]) -> str:
    """Bir çubuğun kesimlerini 'boy ×adet (no)' biçiminde özetler."""
    parts: list[str] = []
    run: list[dict] = []

    def flush():
        if not run:
            return
        length = run[0]["length"]
        ref = run[0].get("element_ref", "")
        count = len(run)
        text = f"{length}" + (f"×{count}" if count > 1 else "")
        if ref:
            text += f" ({ref})"
        parts.append(text)

    for cut in cuts:
        if run and (
            cut["length"] == run[0]["length"]
            and cut.get("element_ref", "") == run[0].get("element_ref", "")
        ):
            run.append(cut)
        else:
            flush()
            run = [cut]
    flush()
    return " + ".join(parts)


def _diameter_summary(result: dict) -> list[dict]:
    """Her çap için metraj/çubuk/fire özetini hesaplar."""
    summary: list[dict] = []
    bar_length_m = float(result.get("bar_length_m", 12.0))

    for diameter, bars in sorted(result.get("plans", {}).items(), key=lambda kv: int(kv[0])):
        total_pieces = sum(len(bar["cuts"]) for bar in bars)
        total_cut_m = sum(cut["length"] for bar in bars for cut in bar["cuts"])
        waste_m = sum(bar["waste_m"] for bar in bars)
        bar_count = len(bars)
        stock_m = bar_count * bar_length_m
        waste_percent = round((waste_m / stock_m) * 100, 2) if stock_m else 0.0
        summary.append(
            {
                "diameter": int(diameter),
                "pieces": total_pieces,
                "cut_length_m": round(total_cut_m, 2),
                "bars": bar_count,
                "stock_length_m": round(stock_m, 2),
                "waste_m": round(waste_m, 2),
                "waste_percent": waste_percent,
            }
        )
    return summary


def build_template() -> io.BytesIO:
    """Donatı girişi için boş XLSX şablonu üretir."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Donati"

    header_fill = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(TEMPLATE_HEADERS, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, example in enumerate(TEMPLATE_EXAMPLES, start=2):
        for col, value in enumerate(example, start=1):
            sheet.cell(row=row_idx, column=col, value=value)

    widths = [12, 12, 10, 16]
    for col, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + col)].width = width

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def export_excel(project_name: str, result: dict) -> io.BytesIO:
    """Optimizasyon sonucunu profesyonel bir Excel raporuna çevirir."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = Workbook()
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(size=14, bold=True)

    # --- Özet sayfası ---
    summary_sheet = workbook.active
    summary_sheet.title = "Özet"
    summary_sheet["A1"] = f"MetrajX — {project_name}"
    summary_sheet["A1"].font = title_font

    headers = ["Çap (mm)", "Parça Adedi", "Kesim Boyu (m)", "Çubuk Sayısı", "Stok Boyu (m)", "Fire (m)", "Fire %"]
    for col, header in enumerate(headers, start=1):
        cell = summary_sheet.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    summary = _diameter_summary(result)
    row = 4
    for item in summary:
        summary_sheet.cell(row=row, column=1, value=f"Ø{item['diameter']}")
        summary_sheet.cell(row=row, column=2, value=item["pieces"])
        summary_sheet.cell(row=row, column=3, value=item["cut_length_m"])
        summary_sheet.cell(row=row, column=4, value=item["bars"])
        summary_sheet.cell(row=row, column=5, value=item["stock_length_m"])
        summary_sheet.cell(row=row, column=6, value=item["waste_m"])
        summary_sheet.cell(row=row, column=7, value=item["waste_percent"])
        row += 1

    summary_sheet.cell(row=row + 1, column=1, value="TOPLAM").font = Font(bold=True)
    summary_sheet.cell(row=row + 1, column=4, value=result.get("total_bars", 0)).font = Font(bold=True)
    summary_sheet.cell(row=row + 1, column=6, value=result.get("total_waste_m", 0)).font = Font(bold=True)
    summary_sheet.cell(row=row + 1, column=7, value=result.get("waste_percent", 0)).font = Font(bold=True)

    for col in range(1, len(headers) + 1):
        summary_sheet.column_dimensions[chr(64 + col)].width = 16

    # --- Kesim planı sayfası ---
    plan_sheet = workbook.create_sheet("Kesim Planı")
    plan_headers = ["Çap (mm)", "Çubuk Adedi", "Kesimler (m) — boy ×adet (eleman no)", "Fire (m)"]
    for col, header in enumerate(plan_headers, start=1):
        cell = plan_sheet.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    plan_row = 2
    bar_length_m = float(result.get("bar_length_m", 12.0))
    for diameter, bars in sorted(result.get("plans", {}).items(), key=lambda kv: int(kv[0])):
        for group in _group_bars(bars):
            bar = group["bar"]
            cuts = _cuts_label(bar["cuts"])
            plan_sheet.cell(row=plan_row, column=1, value=f"Ø{diameter}")
            plan_sheet.cell(row=plan_row, column=2, value=f"×{group['count']}")
            plan_sheet.cell(row=plan_row, column=3, value=f"{cuts}  (/{bar_length_m}m)")
            plan_sheet.cell(row=plan_row, column=4, value=bar["waste_m"])
            plan_row += 1

    plan_sheet.column_dimensions["A"].width = 12
    plan_sheet.column_dimensions["B"].width = 14
    plan_sheet.column_dimensions["C"].width = 70
    plan_sheet.column_dimensions["D"].width = 12

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def export_pdf(project_name: str, result: dict) -> io.BytesIO:
    """Şantiye için görsel kesim çizelgesi (PDF) — site ile uyumlu özet + plan."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.pdfgen import canvas

    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    margin = 1.5 * cm
    content_w = width - 2 * margin
    bar_length_m = float(result.get("bar_length_m", 12.0))

    # Blueprint palette
    C_BG = colors.HexColor("#F3F7FB")
    C_SURFACE = colors.white
    C_PRIMARY = colors.HexColor("#0284C7")
    C_ACCENT = colors.HexColor("#D97706")
    C_FOREGROUND = colors.HexColor("#0C1629")
    C_MUTED = colors.HexColor("#5A6982")
    C_BORDER = colors.HexColor("#D6E0ED")
    C_WASTE = colors.HexColor("#94A3B8")

    total_items = sum(
        len(bar["cuts"]) for bars in result.get("plans", {}).values() for bar in bars
    )
    summary = _diameter_summary(result)

    def draw_page_bg():
        pdf.setFillColor(C_BG)
        pdf.rect(0, 0, width, height, fill=1, stroke=0)

    def draw_header(title: str) -> float:
        draw_page_bg()
        pdf.setFillColor(C_PRIMARY)
        pdf.rect(0, height - 2.4 * cm, width, 2.4 * cm, fill=1, stroke=0)
        pdf.setFillColor(colors.white)
        pdf.setFont("Helvetica-Bold", 15)
        pdf.drawString(margin, height - 1.35 * cm, "MetrajX — Kesim Planı")
        pdf.setFont("Helvetica", 9)
        pdf.drawString(margin, height - 1.85 * cm, title)
        return height - 3.1 * cm

    def draw_metric_box(x: float, y: float, w: float, h: float, label: str, value: str):
        pdf.setFillColor(C_SURFACE)
        pdf.setStrokeColor(C_BORDER)
        pdf.roundRect(x, y, w, h, 5, fill=1, stroke=1)
        pdf.setFillColor(C_MUTED)
        pdf.setFont("Helvetica", 7)
        pdf.drawString(x + 0.35 * cm, y + h - 0.55 * cm, label.upper())
        pdf.setFillColor(C_FOREGROUND)
        pdf.setFont("Helvetica-Bold", 13)
        pdf.drawString(x + 0.35 * cm, y + 0.35 * cm, value)

    def draw_summary_section(y: float) -> float:
        box_w = (content_w - 0.75 * cm) / 4
        box_h = 1.5 * cm
        metrics = [
            ("Fire Oranı", f"%{result.get('waste_percent', 0)}"),
            ("Stok Çubuk", str(result.get("total_bars", 0))),
            ("Çubuk Boyu", f"{bar_length_m} m"),
            ("Donatı Kalemi", str(total_items)),
        ]
        for i, (label, value) in enumerate(metrics):
            draw_metric_box(margin + i * (box_w + 0.25 * cm), y - box_h, box_w, box_h, label, value)
        y -= box_h + 0.6 * cm

        pdf.setFillColor(C_FOREGROUND)
        pdf.setFont("Helvetica-Bold", 10)
        pdf.drawString(margin, y, "Çapa Göre Özet")
        y -= 0.45 * cm

        col_w = [2.2 * cm, 2.2 * cm, 2.5 * cm, 2.5 * cm]
        headers = ["Çap", "Çubuk", "Fire (m)", "Fire %"]
        pdf.setFillColor(C_SURFACE)
        pdf.setStrokeColor(C_BORDER)
        pdf.rect(margin, y - 0.55 * cm, sum(col_w), 0.55 * cm, fill=1, stroke=1)
        pdf.setFillColor(C_MUTED)
        pdf.setFont("Helvetica-Bold", 7)
        cx = margin + 0.2 * cm
        for header, cw in zip(headers, col_w):
            pdf.drawString(cx, y - 0.38 * cm, header)
            cx += cw
        y -= 0.55 * cm

        pdf.setFont("Helvetica", 8)
        for item in summary:
            pdf.setFillColor(C_SURFACE)
            pdf.setStrokeColor(C_BORDER)
            pdf.rect(margin, y - 0.5 * cm, sum(col_w), 0.5 * cm, fill=1, stroke=1)
            pdf.setFillColor(C_FOREGROUND)
            row = [
                f"Ø{item['diameter']}",
                str(item["bars"]),
                str(item["waste_m"]),
                f"%{item['waste_percent']}",
            ]
            cx = margin + 0.2 * cm
            for val, cw in zip(row, col_w):
                pdf.drawString(cx, y - 0.35 * cm, val)
                cx += cw
            y -= 0.5 * cm

        return y - 0.5 * cm

    def draw_cutting_plan(y: float) -> float:
        pdf.setFillColor(C_FOREGROUND)
        pdf.setFont("Helvetica-Bold", 10)
        pdf.drawString(margin, y, "Kesim Planı")
        y -= 0.55 * cm

        bar_height = 0.65 * cm
        track_offset = 1.1 * cm
        track_w = content_w - track_offset - 1.4 * cm

        for diameter, bars in sorted(result.get("plans", {}).items(), key=lambda kv: int(kv[0])):
            if y < 3.5 * cm:
                pdf.showPage()
                y = draw_header(project_name)
                y -= 0.3 * cm

            pdf.setFillColor(C_PRIMARY)
            pdf.setFont("Helvetica-Bold", 9)
            pdf.drawString(margin, y, f"Ø{diameter}  ·  {len(bars)} çubuk")
            y -= 0.55 * cm

            for group in _group_bars(bars):
                bar = group["bar"]
                if y < 2.8 * cm:
                    pdf.showPage()
                    y = draw_header(project_name)
                    y -= 0.3 * cm

                # Count badge
                pdf.setFillColor(colors.HexColor("#E0F2FE"))
                pdf.setStrokeColor(colors.HexColor("#BAE6FD"))
                pdf.roundRect(margin, y - 0.1 * cm, 0.85 * cm, bar_height + 0.1 * cm, 3, fill=1, stroke=1)
                pdf.setFillColor(C_PRIMARY)
                pdf.setFont("Helvetica-Bold", 8)
                pdf.drawCentredString(margin + 0.425 * cm, y + 0.18 * cm, f"×{group['count']}")

                track_x = margin + track_offset
                cursor = track_x

                for cut in bar["cuts"]:
                    seg_w = max((cut["length"] / bar_length_m) * track_w, 0.05 * cm)
                    pdf.setFillColor(C_PRIMARY)
                    pdf.setStrokeColor(colors.HexColor("#0369A1"))
                    pdf.rect(cursor, y, seg_w, bar_height, fill=1, stroke=1)
                    pdf.setFillColor(colors.white)
                    if seg_w > 0.65 * cm:
                        pdf.setFont("Helvetica-Bold", 6.5)
                        pdf.drawCentredString(cursor + seg_w / 2, y + 0.38 * cm, f"{cut['length']}")
                        ref = cut.get("element_ref", "")
                        if ref and seg_w > 1.0 * cm:
                            pdf.setFont("Helvetica", 5.5)
                            pdf.drawCentredString(cursor + seg_w / 2, y + 0.12 * cm, ref[:12])
                    cursor += seg_w

                if bar["waste_m"] > 0:
                    seg_w = max((bar["waste_m"] / bar_length_m) * track_w, 0.05 * cm)
                    pdf.setFillColor(C_WASTE)
                    pdf.setStrokeColor(colors.HexColor("#CBD5E1"))
                    pdf.rect(cursor, y, seg_w, bar_height, fill=1, stroke=1)

                pdf.setFillColor(C_MUTED)
                pdf.setFont("Helvetica", 6.5)
                pdf.drawString(track_x + track_w + 0.15 * cm, y + 0.2 * cm, f"fire {bar['waste_m']}m")
                y -= bar_height + 0.35 * cm

            y -= 0.25 * cm

        return y

    y = draw_header(project_name)
    y = draw_summary_section(y)
    draw_cutting_plan(y)

    pdf.save()
    buffer.seek(0)
    return buffer
