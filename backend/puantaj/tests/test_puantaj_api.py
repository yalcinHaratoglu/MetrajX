from datetime import date
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.test import TestCase
from rest_framework import status
from rest_framework.test import APIClient

from authentication.models import Company
from metraj.models import MetrajCategory, MetrajItem
from puantaj.models import Subcontractor, Timesheet
from sites.models import Site

User = get_user_model()


class PuantajAPITestCase(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.company = Company.objects.create(name="Test İnşaat")
        self.owner = User.objects.create_user(
            email="owner@conmanage.com",
            password="TestPass123!",
            company=self.company,
            role=User.Role.OWNER,
            is_active=True,
        )
        self.site = Site.objects.create(
            company=self.company,
            created_by=self.owner,
            name="A Blok",
            code="A",
        )
        self.category = MetrajCategory.objects.create(
            company=self.company,
            slug="beton",
            name="Beton",
            default_unit="m3",
            is_custom=True,
        )
        self.client.force_authenticate(user=self.owner)

    def test_create_subcontractor_and_timesheet(self):
        sub_resp = self.client.post(
            "/api/puantaj/subcontractors/",
            {
                "site_id": self.site.id,
                "name": "Demir Taşeron",
                "category": self.category.id,
            },
            format="json",
        )
        self.assertEqual(sub_resp.status_code, status.HTTP_201_CREATED)
        sub_id = sub_resp.data["id"]

        ts_resp = self.client.post(
            "/api/puantaj/timesheets/",
            {
                "site_id": self.site.id,
                "subcontractor": sub_id,
                "date": str(date.today()),
                "worker_count": 5,
            },
            format="json",
        )
        self.assertEqual(ts_resp.status_code, status.HTTP_201_CREATED)

        listing = self.client.get(
            f"/api/puantaj/timesheets/?site_id={self.site.id}&year={date.today().year}&month={date.today().month}"
        )
        self.assertEqual(listing.status_code, status.HTTP_200_OK)
        self.assertEqual(len(listing.data), 1)

    def test_hakedis_from_metraj_progress(self):
        sub = Subcontractor.objects.create(
            site=self.site,
            name="Sıvacı Ltd",
            category=self.category,
        )
        MetrajItem.objects.create(
            site=self.site,
            category=self.category,
            subcontractor=sub,
            description="Dış sıva",
            unit="m2",
            quantity=Decimal("100"),
            unit_price=Decimal("250"),
            completion_percent=50,
        )

        response = self.client.get(f"/api/puantaj/hakedis/?site_id={self.site.id}")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(Decimal(response.data["earned_total"]), Decimal("12500"))
        self.assertEqual(Decimal(response.data["contract_total"]), Decimal("25000"))
        self.assertEqual(len(response.data["lines"]), 1)

        sub_hakedis = self.client.get(f"/api/puantaj/subcontractors/{sub.id}/hakedis/")
        self.assertEqual(sub_hakedis.status_code, status.HTTP_200_OK)
        self.assertEqual(Decimal(sub_hakedis.data["earned_total"]), Decimal("12500"))

    def test_settlement_uses_metraj_not_timesheet_for_money(self):
        sub = Subcontractor.objects.create(
            site=self.site,
            name="Kalıpçı",
            category=self.category,
        )
        MetrajItem.objects.create(
            site=self.site,
            category=self.category,
            subcontractor=sub,
            description="Temel",
            unit="m3",
            quantity=Decimal("10"),
            unit_price=Decimal("1000"),
            completion_percent=100,
        )
        Timesheet.objects.create(
            site=self.site,
            subcontractor=sub,
            date=date(2026, 6, 10),
            worker_count=20,
            created_by=self.owner,
        )

        response = self.client.get(
            f"/api/puantaj/settlement/?site_id={self.site.id}&year=2026&month=6"
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["worker_days"], 20)
        self.assertEqual(Decimal(response.data["grand_total"]), Decimal("10000"))

    def test_metraj_item_subcontractor_assignment(self):
        sub = Subcontractor.objects.create(
            site=self.site,
            name="Betoncu",
            category=self.category,
        )
        item_resp = self.client.post(
            "/api/metraj/items/",
            {
                "site_id": self.site.id,
                "category": self.category.id,
                "subcontractor": sub.id,
                "description": "Temel betonu",
                "unit": "m3",
                "quantity": "50",
                "unit_price": "800",
            },
            format="json",
        )
        self.assertEqual(item_resp.status_code, status.HTTP_201_CREATED)
        self.assertEqual(item_resp.data["subcontractor"], sub.id)
        self.assertEqual(item_resp.data["subcontractor_name"], "Betoncu")

    def test_duplicate_timesheet_same_day_rejected(self):
        sub = Subcontractor.objects.create(
            site=self.site,
            name="Kalıpçı",
            category=self.category,
        )
        payload = {
            "site_id": self.site.id,
            "subcontractor": sub.id,
            "date": "2026-06-15",
            "worker_count": 3,
        }
        first = self.client.post("/api/puantaj/timesheets/", payload, format="json")
        self.assertEqual(first.status_code, status.HTTP_201_CREATED)
        second = self.client.post("/api/puantaj/timesheets/", payload, format="json")
        self.assertEqual(second.status_code, status.HTTP_400_BAD_REQUEST)

    def test_attendance_matrix_xlsx_export(self):
        from puantaj.models import Worker

        sub = Subcontractor.objects.create(
            site=self.site,
            name="Demir Taşeron",
            category=self.category,
        )
        worker = Worker.objects.create(
            subcontractor=sub,
            first_name="Ali",
            last_name="Yılmaz",
        )
        Timesheet.objects.create(
            site=self.site,
            subcontractor=sub,
            worker=worker,
            date=date(2026, 6, 10),
            worker_count=1,
            created_by=self.owner,
        )

        response = self.client.get(
            f"/api/puantaj/attendance-matrix/?site_id={self.site.id}"
            f"&date_from=2026-06-01&date_to=2026-06-30&export=xlsx"
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("spreadsheetml", response["Content-Type"])
        from io import BytesIO

        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(response.content))
        rows = list(wb.active.iter_rows(values_only=True))
        flat = " ".join(str(cell) for row in rows for cell in row if cell is not None)
        self.assertIn("Ali Yılmaz", flat)
        self.assertIn("Demir Taşeron", flat)

    def test_direct_worker_attendance(self):
        from puantaj.models import Worker

        guard = Worker.objects.create(
            site=self.site,
            employment_type=Worker.EmploymentType.DIRECT,
            role=Worker.Role.SECURITY,
            pay_type=Worker.PayType.MONTHLY,
            first_name="Mehmet",
            last_name="Bekçi",
        )
        self.client.force_authenticate(user=self.owner)
        matrix = self.client.get(
            f"/api/puantaj/attendance-matrix/?site_id={self.site.id}"
            f"&date_from=2026-06-01&date_to=2026-06-30&employment_type=direct"
        )
        self.assertEqual(matrix.status_code, status.HTTP_200_OK)
        names = [row["full_name"] for row in matrix.data["workers"]]
        self.assertIn("Mehmet Bekçi", names)

        toggle = self.client.post(
            "/api/puantaj/attendance-toggle/",
            {
                "site_id": self.site.id,
                "worker_id": guard.id,
                "date": "2026-06-12",
                "present": True,
            },
            format="json",
        )
        self.assertEqual(toggle.status_code, status.HTTP_200_OK)
        self.assertTrue(
            Timesheet.objects.filter(worker=guard, date=date(2026, 6, 12)).exists()
        )

    def test_create_direct_worker(self):
        from puantaj.models import Worker

        self.client.force_authenticate(user=self.owner)
        response = self.client.post(
            "/api/puantaj/workers/",
            {
                "site_id": self.site.id,
                "employment_type": "direct",
                "role": "security_guard",
                "pay_type": "monthly",
                "first_name": "Ayşe",
                "last_name": "Kapıcı",
            },
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        worker = Worker.objects.get(pk=response.data["id"])
        self.assertEqual(worker.employment_type, Worker.EmploymentType.DIRECT)
        self.assertIsNone(worker.subcontractor_id)
        self.assertEqual(worker.site_id, self.site.id)
